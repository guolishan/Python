# -*- coding: utf-8 -*-

import sys
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication, QMainWindow,QFileDialog
import os

import requests
from fake_useragent import UserAgent
import json
import  logging
import time
import pandas as pd
from openpyxl import load_workbook

class Ui_Form(object):
    def setupUi(self, Form):
        Form.setObjectName("Form")
        Form.resize(796, 519)
        self.label_4 = QtWidgets.QLabel(Form)
        self.label_4.setGeometry(QtCore.QRect(120, 170, 29, 20))
        self.label_4.setObjectName("label_4")
        self.rizhi_textBrowser = QtWidgets.QTextBrowser(Form)
        self.rizhi_textBrowser.setGeometry(QtCore.QRect(170, 190, 451, 311))
        self.rizhi_textBrowser.setObjectName("rizhi_textBrowser")
        self.label_3 = QtWidgets.QLabel(Form)
        self.label_3.setGeometry(QtCore.QRect(249, 60, 101, 31))
        font = QtGui.QFont()
        font.setFamily("微软雅黑")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.Typelist_comboBox = QtWidgets.QComboBox(Form)
        self.Typelist_comboBox.setGeometry(QtCore.QRect(359, 60, 260, 30))
        self.Typelist_comboBox.setMinimumSize(QtCore.QSize(200, 30))
        font = QtGui.QFont()
        font.setFamily("微软雅黑")
        font.setPointSize(10)
        self.Typelist_comboBox.setFont(font)
        self.Typelist_comboBox.setObjectName("Typelist_comboBox")
        self.dataout_pushButton = QtWidgets.QPushButton(Form)
        self.dataout_pushButton.setGeometry(QtCore.QRect(626, 60, 75, 71))
        self.dataout_pushButton.setMinimumSize(QtCore.QSize(75, 30))
        font = QtGui.QFont()
        font.setFamily("微软雅黑")
        font.setPointSize(10)
        font.setBold(True)
        font.setItalic(False)
        font.setUnderline(False)
        font.setWeight(75)
        font.setStrikeOut(False)
        self.dataout_pushButton.setFont(font)
        self.dataout_pushButton.setAutoDefault(False)
        self.dataout_pushButton.setDefault(True)
        self.dataout_pushButton.setFlat(False)
        self.dataout_pushButton.setObjectName("dataout_pushButton")
        self.filepath_lineEdit = QtWidgets.QLineEdit(Form)
        self.filepath_lineEdit.setGeometry(QtCore.QRect(249, 150, 370, 30))
        self.filepath_lineEdit.setMinimumSize(QtCore.QSize(370, 30))
        font = QtGui.QFont()
        font.setFamily("微软雅黑")
        font.setPointSize(10)
        self.filepath_lineEdit.setFont(font)
        self.filepath_lineEdit.setObjectName("filepath_lineEdit")
        self.file_pushButton = QtWidgets.QPushButton(Form)
        self.file_pushButton.setGeometry(QtCore.QRect(626, 150, 75, 30))
        self.file_pushButton.setMinimumSize(QtCore.QSize(75, 30))
        font = QtGui.QFont()
        font.setFamily("微软雅黑")
        font.setPointSize(10)
        font.setBold(True)
        font.setUnderline(True)
        font.setWeight(75)
        self.file_pushButton.setFont(font)
        self.file_pushButton.setObjectName("file_pushButton")
        self.layoutWidget = QtWidgets.QWidget(Form)
        self.layoutWidget.setGeometry(QtCore.QRect(110, 60, 135, 32))
        self.layoutWidget.setObjectName("layoutWidget")
        self.formLayout = QtWidgets.QFormLayout(self.layoutWidget)
        self.formLayout.setContentsMargins(0, 0, 0, 0)
        self.formLayout.setObjectName("formLayout")
        self.label = QtWidgets.QLabel(self.layoutWidget)
        font = QtGui.QFont()
        font.setFamily("微软雅黑")
        font.setPointSize(10)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.formLayout.setWidget(0, QtWidgets.QFormLayout.LabelRole, self.label)
        self.start_dateEdit = QtWidgets.QDateEdit(self.layoutWidget)
        self.start_dateEdit.setMinimumSize(QtCore.QSize(75, 30))
        font = QtGui.QFont()
        font.setFamily("微软雅黑")
        font.setPointSize(10)
        self.start_dateEdit.setFont(font)
        self.start_dateEdit.setDateTime(QtCore.QDateTime(QtCore.QDate(2017, 12, 31), QtCore.QTime(0, 0, 0)))
        self.start_dateEdit.setDate(QtCore.QDate(2017, 12, 31))
        self.start_dateEdit.setObjectName("start_dateEdit")
        self.formLayout.setWidget(0, QtWidgets.QFormLayout.FieldRole, self.start_dateEdit)
        self.layoutWidget1 = QtWidgets.QWidget(Form)
        self.layoutWidget1.setGeometry(QtCore.QRect(110, 98, 135, 32))
        self.layoutWidget1.setObjectName("layoutWidget1")
        self.formLayout_2 = QtWidgets.QFormLayout(self.layoutWidget1)
        self.formLayout_2.setContentsMargins(0, 0, 0, 0)
        self.formLayout_2.setObjectName("formLayout_2")
        self.label_2 = QtWidgets.QLabel(self.layoutWidget1)
        font = QtGui.QFont()
        font.setFamily("微软雅黑")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.formLayout_2.setWidget(0, QtWidgets.QFormLayout.LabelRole, self.label_2)
        self.end_dateEdit = QtWidgets.QDateEdit(self.layoutWidget1)
        self.end_dateEdit.setMinimumSize(QtCore.QSize(75, 30))
        font = QtGui.QFont()
        font.setFamily("微软雅黑")
        font.setPointSize(10)
        self.end_dateEdit.setFont(font)
        self.end_dateEdit.setDateTime(QtCore.QDateTime(QtCore.QDate(2019, 12, 31), QtCore.QTime(0, 0, 0)))
        self.end_dateEdit.setMaximumDate(QtCore.QDate(9999, 12, 31))
        self.end_dateEdit.setObjectName("end_dateEdit")
        self.formLayout_2.setWidget(0, QtWidgets.QFormLayout.FieldRole, self.end_dateEdit)
        self.label_5 = QtWidgets.QLabel(Form)
        self.label_5.setGeometry(QtCore.QRect(250, 100, 101, 31))
        font = QtGui.QFont()
        font.setFamily("微软雅黑")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")
        self.lineEditCode = QtWidgets.QLineEdit(Form)
        self.lineEditCode.setGeometry(QtCore.QRect(360, 99, 260, 31))
        self.lineEditCode.setMinimumSize(QtCore.QSize(200, 0))
        self.lineEditCode.setObjectName("lineEditCode")
        self.widget = QtWidgets.QWidget(Form)
        self.widget.setGeometry(QtCore.QRect(140, 10, 614, 32))
        self.widget.setMinimumSize(QtCore.QSize(200, 30))
        self.widget.setObjectName("widget")
        self.horizontalLayout = QtWidgets.QHBoxLayout(self.widget)
        self.horizontalLayout.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.radioButtonCN = QtWidgets.QRadioButton(self.widget)
        self.radioButtonCN.setMinimumSize(QtCore.QSize(200, 30))
        font = QtGui.QFont()
        font.setFamily("微软雅黑")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.radioButtonCN.setFont(font)
        self.radioButtonCN.setObjectName("radioButtonCN")
        self.horizontalLayout.addWidget(self.radioButtonCN)
        self.radioButtonUS = QtWidgets.QRadioButton(self.widget)
        self.radioButtonUS.setMinimumSize(QtCore.QSize(200, 30))
        font = QtGui.QFont()
        font.setFamily("微软雅黑")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.radioButtonUS.setFont(font)
        self.radioButtonUS.setObjectName("radioButtonUS")
        self.horizontalLayout.addWidget(self.radioButtonUS)
        self.radioButtonHK = QtWidgets.QRadioButton(self.widget)
        self.radioButtonHK.setMinimumSize(QtCore.QSize(200, 30))
        font = QtGui.QFont()
        font.setFamily("微软雅黑")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.radioButtonHK.setFont(font)
        self.radioButtonHK.setObjectName("radioButtonHK")
        self.horizontalLayout.addWidget(self.radioButtonHK)

        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "雪球网数据提取小工具——by：才哥"))
        self.label_4.setText(_translate("Form", "<html><head/><body><p align=\"center\"><span style=\" font-size:10pt; font-weight:600; font-style:italic;\">日志</span></p></body></html>"))
        self.label_3.setText(_translate("Form", "<html><head/><body><p align=\"center\">请选择数据类型</p></body></html>"))
        self.dataout_pushButton.setText(_translate("Form", "导出数据"))
        self.filepath_lineEdit.setToolTip(_translate("Form", "<html><head/><body><p><br/></p></body></html>"))
        self.filepath_lineEdit.setText(_translate("Form", "请选择你要导出文件的目录"))
        self.file_pushButton.setText(_translate("Form", "选择文件夹"))
        self.label.setText(_translate("Form", "<html><head/><body><p align=\"center\"><span style=\" font-size:10pt; font-weight:600; text-decoration: underline; color:#ffaa00;\">开始年份</span></p></body></html>"))
        self.start_dateEdit.setDisplayFormat(_translate("Form", "yyyy"))
        self.label_2.setText(_translate("Form", "<html><head/><body><p align=\"center\"><span style=\" font-size:10pt; font-weight:600; text-decoration: underline; color:#ffaa00;\">截止年份</span></p></body></html>"))
        self.end_dateEdit.setDisplayFormat(_translate("Form", "yyyy"))
        self.label_5.setText(_translate("Form", "<html><head/><body><p align=\"center\">请输入股票代码</p></body></html>"))
        self.radioButtonCN.setText(_translate("Form", "A股"))
        self.radioButtonUS.setText(_translate("Form", "美股"))
        self.radioButtonHK.setText(_translate("Form", "港股"))

class MyMainForm(QMainWindow, Ui_Form):
    def __init__(self, parent=None):
        super(MyMainForm, self).__init__()
        self.setupUi(self)
        log = '股票代码列表已刷新\n请选择股票代码、数据类型及存储目录文件夹后进行进行 数据导出\n 股票代码中输入 全选 可以导出全部，容易死机，谨慎使用'
        self.rizhi_textBrowser.append(log)
        #self.
        self.Typelist_comboBox.addItems(list(dataType.keys())) #数据报告类型选择
        # self.codelist_comboBox.addItems(list(ipoCode.keys())) #股票代码选择
#        self.textBrowser.setText('点击转换获货币兑换关系')
        self.dataout_pushButton.clicked.connect(self.Get_data) #获取并导出数据
        self.file_pushButton.clicked.connect(self.openFile)

    def openFile(self):
        get_directory_path = QFileDialog.getExistingDirectory(self,
                                    "选取指定文件夹",
                                    "C:/")
        self.filepath_lineEdit.setText(str(get_directory_path))
    
    # def dataOut(self):
        
    #     self.textBrowser.setText(str_print)
    
    # def AB_type(self):
    #             #股市类型
    #     if (self.radioButtonCN.isChecked()):
    #         ABtype = 'cn'
    #     elif (self.radioButtonUS.isChecked()):
    #         ABtype = 'us'
    #     elif (self.radioButtonHK.isChecked()):
    #         ABtype = 'hk'
    #     else:
    #         ABtype = 'cn'
            
    #     return ABtype

    def Get_url(self,name,ipo_code):
        #获取开始结束时间戳（开始和结束时间手动输入）
        inputstartTime = str(self.start_dateEdit.date().toPyDate().year)
        inputendTime = str(self.end_dateEdit.date().toPyDate().year)
        endTime = f'{inputendTime}-12-31 00:00:00'
        timeArray = time.strptime(endTime, "%Y-%m-%d %H:%M:%S")

        #获取指定的数据类型及股票代码
        ipo_code = ipo_code
        filename = ipo_code
        name = name
        data_type =dataType[name]
        #计算需要采集的数据量(一年以四个算)
        count_num = (int(inputendTime) - int(inputstartTime) +1) * 4
        start_time =  f'{int(time.mktime(timeArray))}001'
        
        #股市类型
        if (self.radioButtonCN.isChecked()):
            ABtype = 'cn'
            num = 3
        elif (self.radioButtonUS.isChecked()):
            ABtype = 'us'
            num = 6
        elif (self.radioButtonHK.isChecked()):
            ABtype = 'hk'
            num = 6
        else:
            ABtype = 'cn'
            num = 3
        
        #基础网站
        base_url = f'https://stock.xueqiu.com/v5/stock/finance/{ABtype}'

        #组合url地址
        url = f'{base_url}/{data_type}.json?symbol={ipo_code}&type=all&is_detail=true&count={count_num}&timestamp={start_time}'
        logging.info('获取%s的%s详情页url为 %s',filename,name,url)
        log = f'{filename}的{name}数据网址已经解析成功'
        self.rizhi_textBrowser.append(log)
        return url,num

    #获取原始数据
    def Get_data(self):
        name = self.Typelist_comboBox.currentText()      
        ipo_code = self.lineEditCode.text()
        if (self.radioButtonCN.isChecked()):
            ipoCodex=ipoCodecn
        elif (self.radioButtonUS.isChecked()):
            ipoCodex=ipoCodeus
        elif (self.radioButtonHK.isChecked()):
            ipoCodex=ipoCodehk
        else:
            ipoCodex=ipoCodecn

        if name == '全选' and ipo_code == '全选':
            for ipo_code in list(ipoCodex.keys()):
                for name in list(dataType.keys())[1:]:
                    self.re_data(name,ipo_code)
        elif name == '全选' and ipo_code != '全选':
                for name in list(dataType.keys())[1:]:
                    self.re_data(name,ipo_code)
        elif ipo_code == '全选' and name != '全选':
            for ipo_code in list(ipoCodex.keys()):
                self.re_data(name,ipo_code)            
        else:
            self.re_data(name,ipo_code)
    
    def re_data(self,name,ipo_code):
        filepath = self.filepath_lineEdit.text()
        filename = ipoCode[ipo_code]
        name = name
        url,num = self.Get_url(name,ipo_code)
        headers = {"User-Agent": UserAgent(verify_ssl=False).random}
        df = requests.get(url,headers = headers,cookies = cookies)
  
        df = df.text

        try:
            data = json.loads(df)
            pd_df = pd.DataFrame(data['data']['list'])
                    
            path = f'{filepath}\{filename}.xlsx'    
            cols = pd_df.columns.tolist()
            data = pd.DataFrame()            
            
                       
            data['报告名称'] = pd_df['report_name']
            for i in range(num,len(cols)):
                col = cols[i]
                try:
                    data[col] = pd_df[col].apply(lambda x:x[0])
                # data[f'{col}_同比'] = pd_df[col].apply(lambda x:x[1])
                except TypeError:
                    pass
            data = data.set_index('报告名称')       
            logging.info('%s的%s数据已经爬取成功',filename,name)
            log = f'{filename}的{name}数据已经爬取成功'
            self.rizhi_textBrowser.append(log)
            dataT = data.T
            dataT.rename(index = eval(f'_{name}'),inplace=True)
            try:
                if os.path.exists(path):
                    df_dic = pd.read_excel(path,None)
                    if name not in list(df_dic.keys()):
                        logging.info('%s的%s数据页签不存在',filename,name)
                        log = f'{filename}的{name}数据页签不存在，创建新页签'
                        self.rizhi_textBrowser.append(log)
                        with pd.ExcelWriter(path,mode='a') as writer:
                            book = load_workbook(path)     
                            writer.book = book    
                            dataT.to_excel(writer,sheet_name=name)
                            writer.save()
                    else:
                        logging.info('%s的%s数据已存在',filename,name)
                        log = f'{filename}的{name}数据页签已存在，合并中'
                        self.rizhi_textBrowser.append(log)
                        df = pd.read_excel(path,sheet_name = name,index_col=0)
                        d_ = list(set(list(dataT.columns)) - set(list(df.columns)))
                        logging.info('读取原页签')
                        dataT = pd.merge(df,dataT[d_],how='outer',left_index=True,right_index=True)
                        dataT.sort_index(axis=1,ascending=False,inplace=True)
                        with pd.ExcelWriter(path,engine='openpyxl') as writer:   
                            book = load_workbook(path)     
                            writer.book = book
                            idx = writer.book.sheetnames.index(name)
                            writer.book.remove(writer.book.worksheets[idx])
                            writer.book.create_sheet(name, idx)
                            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}         
                            
                            dataT.to_excel(writer,sheet_name=name,startcol=0)
                            writer.save()
                else:
                    dataT.to_excel(path,sheet_name=name)
                    
                logging.info('%s的%s数据已经保存成功',filename,name)
                log = f'<font color=\"#00CD00\">{filename}的{name}数据已经保存成功</font>'
                self.rizhi_textBrowser.append(log)
            except FileNotFoundError:
                logging.info('未设置存储目录或存储目录不存在')
                log = '<font color=\"#FF0000\">未设置存储目录或存储目录不存在,请重新选择文件夹</font>'
                self.rizhi_textBrowser.append(log)
        except KeyError:
            logging.info('该股票此类型报告不存在')
            log = '<font color=\"#FF0000\">该股票此类型报告不存在,请重新选择股票代码或数据类型</font>'
            self.rizhi_textBrowser.append(log)              
                
#自动获取雪球网cookies，用于后续采集数据时自动记录访问状态
def Get_Cookie(url):
    session = requests.Session()
    headers = {"User-Agent": UserAgent(verify_ssl=False).random}
   
    session.get(url, headers=headers)
    
    #获取当前的Cookie
    Cookie= dict(session.cookies)
    
    return Cookie

#获取上市公司列表（股票代号和名称）
def Get_IPOList(cookies):
    headers = {"User-Agent": UserAgent(verify_ssl=False).random}
    url = 'https://xueqiu.com/service/v5/stock/screener/quote/list?page=1&size=9999&order=desc&orderby=percent&order_by=percent&market=CN&type=sh_sz'
    url1 = 'https://xueqiu.com/service/v5/stock/screener/quote/list?page=1&size=9999&order=desc&orderby=percent&order_by=percent&market=US&type=us'
    url2 = 'https://xueqiu.com/service/v5/stock/screener/quote/list?page=1&size=9999&order=desc&orderby=percent&order_by=percent&market=HK&type=hk'
    response = requests.get(url,headers = headers,cookies = cookies)
    response1 = requests.get(url1,headers = headers,cookies = cookies)
    response2 = requests.get(url2,headers = headers,cookies = cookies)
    
    df = response.text
    data = json.loads(df)   
    data = data['data']['list']
    data = pd.DataFrame(data)
    data = data[['symbol','name']]
    data['name'] = data['symbol']+' '+data['name']
    data.sort_values(by = ['symbol'],inplace=True)
    data = data.set_index(data['symbol'])['name']
    ipoCodecn = data.to_dict()

    df1 = response1.text
    data1 = json.loads(df1)    
    data1 = data1['data']['list']
    data1 = pd.DataFrame(data1)
    data1 = data1[['symbol','name']]
    data1['name'] = data1['symbol']+' '+data1['name']
    data1.sort_values(by = ['symbol'],inplace=True)
    data1 = data1.set_index(data1['symbol'])['name']
    ipoCodeus = data1.to_dict()

    df2 = response2.text
    data2 = json.loads(df2)
    data2 = data2['data']['list']
    data2 = pd.DataFrame(data2)
    data2 = data2[['symbol','name']]
    data2['name'] = data2['symbol']+' '+data2['name']
    data2.sort_values(by = ['symbol'],inplace=True)
    data2 = data2.set_index(data2['symbol'])['name']
    ipoCodehk = data2.to_dict()

    xx = dict(ipoCodecn,**ipoCodehk)
    ipoCode = dict(xx,**ipoCodeus)
    # fileList = [data,data1,data2]
    # data = pd.concat(fileList,ignore_index = True)
    
    # data.sort_values(by = ['symbol'],inplace=True)
    # data = data.set_index(data['symbol'])['name']
    # ipoCode = data.to_dict()
    
    return ipoCode,ipoCodecn,ipoCodehk,ipoCodeus
      

if __name__ == "__main__":
    
    logging.basicConfig(level = logging.INFO,
                    format = '%(asctime)s-%(levelname)s:%(message)s')
    logging.info('正在加载程序...........')
    original_url = 'https://xueqiu.com'
    #数据类型字典
    dataType = {'全选':'all',
                '主要指标':'indicator',
                '利润表':'income',
                '资产负债表':'balance',
                '现金流量表':'cash_flow'}
    _主要指标 ={'account_receivable_turnover': '应收账款周转率', 'accounts_payable_turnover': '应付账款周转率', 'accounts_payable_turnover_days': '应付账款周转天数', 'asset_liab_ratio': '资产负债率', 'avg_roe': '净资产收益率', 'basic_eps': '每股收益', 'capital_reserve': '每股资本公积金', 'cash_cycle': '现金循环周期', 'current_asset_turnover_rate': '流动资产周转率', 'current_ratio': '流动比率', 'equity_multiplier': '权益乘数', 'equity_ratio': '产权比率', 'fixed_asset_turnover_ratio': '固定资产周转率', 'gross_selling_rate': '销售毛利率', 'holder_equity': '股东权益比率', 'inventory_turnover': '存货周转率', 'inventory_turnover_days': '存货周转天数', 'ncf_from_oa_to_total_liab': '现金流量比率', 'net_interest_of_total_assets': '总资产报酬率', 'net_profit_after_nrgal_atsolc': '扣非净利润', 'net_profit_atsopc': '净利润', 'net_profit_atsopc_yoy': '净利润同比增长', 'net_selling_rate': '销售净利率', 'np_atsopc_nrgal_yoy': '扣非净利润同比增长', 'np_per_share': '每股净资产', 'operate_cash_flow_ps': '每股经营现金流', 'operating_cycle': '营业周期', 'operating_income_yoy': '营业收入同比增长', 'ore_dlt': '净资产收益率-摊薄', 'quick_ratio': '速动比率', 'receivable_turnover_days': '应收账款周转天数', 'rop': '人力投入回报率', 'total_capital_turnover': '总资产周转率', 'total_revenue': '营业收入', 'undistri_profit_ps': '每股未分配利润', 'advance_premium': '预收保费', 'amortized_cost_fnncl_assets': '以摊余成本计量的金融资产', 'assured_pledge_loan': '保户质押贷款', 'assured_saving_and_invest': '保户储金及投资款', 'borrowing_funds': '拆入资金', 'charge_and_commi_payable': '应付手续费及佣金', 'claim_payable': '应付赔付款', 'derivative_fnncl_assets': '衍生金融资产', 'disbursement_loan_and_advance': '发放贷款和垫款', 'dvdnd_payable_for_the_insured': '应付保单红利', 'fixed_deposit': '定期存款', 'fnncl_assets_sold_for_repur': '卖出回购金融资产款', 'fv_chg_income_fnncl_assets': '以公允价值计量且其变动计入其他综合收益的金融资产', 'independent_account_liab': '独立账户负债', 'insurance_contract_reserve': '保险合同准备金', 'interbank_deposit_etc': '同业及其他金融机构存放款项', 'lending_fund': '拆出资金', 'life_insurance_reserve': '寿险责任准备金', 'lt_health_insurance_reserve': '长期健康险责任准备金', 'othr_assets': '其他资产', 'othr_liab': '其他负债', 'paid_capital_deposit': '存出资本保证金', 'premium_receivable': '应收保费', 'receivable_rein_duty_reserve': '应收分保寿险责任准备金', 'received_deposit': '存入保证金', 'rein_account_receivable': '应收分保账款', 'rein_contract_reserve': '应收分保合同准备金', 'rein_payable': '应付分保账款', 'rein_undue_liability_reserve': '应收分保未到期责任准备金', 'separate_account': '独立账户资产', 'unearned_premium_reserve': '应收分保未到期责任准备金'}
    _利润表 = {'amortized_deposit_for_duty': '减：摊回保险责任准备金', 'amortized_rein_expenditure': '减：摊回分保费用', 'asset_disposal_income': '资产处置收益', 'asset_impairment_loss': '资产减值损失', 'basic_eps': '基本每股收益', 'business_and_manage_fee': '业务及管理费', 'ceded_out_premium': '减：分出保费', 'charge_and_commi_expenses': '手续费及佣金支出', 'commi_on_insurance_policy': '保单红利支出', 'compen_expense': '减：摊回赔付支出', 'compen_payout': '赔付支出', 'continous_operating_np': '（一）持续经营净利润', 'credit_impairment_loss': '信用减值损失', 'dlt_earnings_per_share': '稀释每股收益', 'draw_duty_deposit': '提取保险责任准备金', 'draw_undueduty_deposit': '提取未到期责任准备金', 'earned_premium': '已赚保费', 'exchg_gain': '汇兑收益', 'finance_cost_interest_fee': '其中：利息费用', 'finance_cost_interest_income': '利息收入', 'financing_expenses': '财务费用', 'income_from_chg_in_fv': '加：公允价值变动收益 ', 'income_tax_expenses': '减：所得税费用', 'insurance_income': '保险业务收入', 'invest_income': '投资收益', 'invest_incomes_from_rr': '其中：对联营企业和合营企业的投资收益', 'manage_fee': '管理费用', 'minority_gal': '少数股东损益', 'net_profit': '净利润', 'net_profit_after_nrgal_atsolc': '扣除非经常性损益后的净利润', 'net_profit_atsopc': '归属于母公司所有者的净利润', 'net_profit_bi': '净利润差额(合计平衡项目)', 'non_operating_income': '加：营业外收入', 'non_operating_payout': '减：营业外支出', 'noncurrent_asset_disposal_loss': '其中：非流动资产处置损失', 'noncurrent_assets_dispose_gain': '其中：非流动资产处置利得', 'op': '营业利润', 'operating_cost': '其中：营业成本', 'operating_costs': '营业总成本', 'operating_payout': '营业总成本', 'operating_taxes_and_surcharge': '营业税金及附加', 'other_income': '其他收益', 'othr_business_costs': '其他业务成本', 'othr_compre_income': '其他综合收益', 'othr_compre_income_atms': '归属于少数股东的其他综合收益', 'othr_compre_income_atoopc': '归属母公司所有者的其他综合收益', 'othr_income': '其他业务收入', 'profit_total_amt': '利润总额', 'rad_cost': '研发费用', 'refunded_premium': '退保金', 'rein_expenditure': '分保费用', 'rein_premium_income': '其中：分保费收入', 'revenue': '其中：营业收入', 'sales_fee': '销售费用', 'total_compre_income': '综合收益总额', 'total_compre_income_atms': '归属于少数股东的综合收益总额', 'total_compre_income_atsopc': '归属于母公司股东的综合收益总额', 'total_revenue': '营业中收入'}
    _现金流量表 = {'branch_paid_to_minority_holder': '其中：子公司支付给少数股东的股利', 'cash_paid_for_assets': '购建固定资产、无形资产和其他长期资产支付的现金', 'cash_paid_of_distribution': '分配股利、利润或偿付利息支付的现金', 'cash_paid_to_employee_etc': '支付给职工以及为职工支付的现金', 'cash_pay_for_debt': '偿还债务支付的现金', 'cash_received_from_bond_issue': '发行债券收到的现金', 'cash_received_from_investor': '其中：子公司吸收少数股东投资收到的现金', 'cash_received_of_absorb_invest': '吸收投资收到的现金', 'cash_received_of_borrowing': '取得借款收到的现金', 'cash_received_of_dspsl_invest': '收回投资收到的现金', 'cash_received_of_othr_fa': '收到其他与筹资活动有关的现金', 'cash_received_of_othr_ia': '收到其他与投资活动有关的现金', 'cash_received_of_othr_oa': '收到其他与经营活动有关的现金', 'cash_received_of_sales_service': '销售商品、提供劳务收到的现金', 'effect_of_exchange_chg_on_cce': '汇率变动对现金及现金等价物的影响', 'final_balance_of_cce': '期末现金及现金等价物余额', 'goods_buy_and_service_cash_pay': '购买商品、接受劳务支付的现金', 'initial_balance_of_cce': '加：期初现金及现金等价物余额', 'invest_income_cash_received': '取得投资收益收到的现金', 'invest_paid_cash': '投资支付的现金', 'ncf_from_fa': '筹资活动产生的现金流量净额', 'ncf_from_ia': '投资活动产生的现金流量净额', 'ncf_from_oa': '经营活动产生的现金流量净额', 'net_cash_amt_from_branch': '取得子公司及其他营业单位支付的现金净额', 'net_cash_of_disposal_assets': '处置固定资产、无形资产和其他长期资产收回的现金净额', 'net_cash_of_disposal_branch': '处置子公司及其他营业单位收到的现金净额', 'net_increase_in_cce': '现金及现金等价物净增加额', 'othrcash_paid_relating_to_fa': '支付其他与筹资活动有关的现金', 'othrcash_paid_relating_to_ia': '支付其他与投资活动有关的现金', 'othrcash_paid_relating_to_oa': '支付其他与经营活动有关的现金', 'payments_of_all_taxes': '支付的各项税费', 'refund_of_tax_and_levies': '收到的税费返还', 'sub_total_of_ci_from_fa': '筹资活动现金流入小计', 'sub_total_of_ci_from_ia': '投资活动现金流入小计', 'sub_total_of_ci_from_oa': '经营活动现金流入小计', 'sub_total_of_cos_from_fa': '筹资活动现金流出小计', 'sub_total_of_cos_from_ia': '投资活动现金流出小计', 'sub_total_of_cos_from_oa': '经营活动现金流出小计', 'cash_received_from_orig_ic': '收到原保险合同保费取得的现金', 'net_cash_received_from_rein': '收到再保业务现金净额', 'naa_assured_saving_and_invest': '保户储金及投资款净增加额', 'oa_net_ci_si': '收到的税费返还', 'cash_of_orig_ic_indemnity': '支付原保险合同赔付款项的现金', 'oa_net_cos_si': '支付再保险业务现金净额', 'cash_paid_for_fees_and_commi': '支付利息、手续费及佣金的现金', 'cash_paid_for_policy_dividends': '支付保单红利的现金', 'net_increase_in_pledge_loans': '质押贷款净增加额', 'ia_cos_si': '取得子公司及其他营业单位支付的现金净额'}
    _资产负债表 ={'total_assets': '资产合计', 'total_liab_and_holders_equity': '负债和股东权益总计', 'total_holders_equity': '股东权益合计', 'total_quity_atsopc': '归属于母公司股东权益合计', 'total_noncurrent_assets': '非流动资产合计', 'total_current_assets': '流动资产合计', 'total_liab': '负债合计', 'fixed_asset_sum': '固定资产合计', 'undstrbtd_profit': '未分配利润', 'total_current_liab': '流动负债合计', 'inventory': '存货', 'total_noncurrent_liab': '非流动负债合计', 'account_receivable': '应收票据及应收账款', 'ar_and_br': '应收账款', 'lt_loan': '长期借款', 'shares': '实收资本(或股本)', 'currency_funds': '货币资金', 'capital_reserve': '资本公积', 'st_loan': '短期借款', 'construction_in_process_sum': '在建工程合计', 'intangible_assets': '无形资产', 'bp_and_ap': '应付票据及应付账款', 'earned_surplus': '盈余公积', 'noncurrent_liab_due_in1y': '一年内到期的非流动负债', 'accounts_payable': '应付账款', 'lt_equity_invest': '长期股权投资', 'dev_expenditure': '开发支出', 'noncurrent_liab_di': '递延收益-非流动负债', 'minority_equity': '少数股东权益', 'bill_payable': '应付票据', 'other_illiquid_fnncl_assets': '其他非流动金融资产', 'dt_assets': '递延所得税资产', 'estimated_liab': '预计负债', 'payroll_payable': '应付职工薪酬', 'tax_payable': '应交税费', 'othr_current_assets': '其他流动资产', 'pre_receivable': '预收款项', 'goodwill': '商誉', 'pre_payment': '预付款项', 'othr_compre_income': '其他综合收益', 'othr_noncurrent_assets': '其他非流动资产', 'lt_deferred_expense': '长期待摊费用', 'other_eq_ins_invest': '其他权益工具投资', 'othr_non_current_liab': '其他非流动负债', 'invest_property': '投资性房地产', 'dt_liab': '递延所得税负债', 'asset_liab_ratio': '资产负债比', 'lt_payable_sum': '长期应付款合计', 'bills_receivable': '其中：应收票据', 'bond_payable': '应付债券', 'construction_in_process': '其中：在建工程', 'contract_liabilities': '合同负债', 'contractual_assets': '合同资产', 'current_assets_si': '流动资产', 'current_liab_si': '流动负债', 'derivative_fnncl_liab': '衍生金融负债', 'dividend_payable': '应付股利', 'dividend_receivable': '应收股利', 'fixed_asset': '固定资产', 'fixed_assets_disposal': '固定资产清理', 'frgn_currency_convert_diff': '外币报表折算差额', 'general_risk_provision': '一般风险准备', 'held_to_maturity_invest': '持有至到期投资', 'interest_payable': '应付利息', 'interest_receivable': '应收利息', 'lt_payable': '长期应付款', 'lt_receivable': '长期应收款', 'nca_due_within_one_year': '一年内到期的非流动资产', 'noncurrent_assets_si': '非流动资产', 'noncurrent_liab_si': '非流动负债', 'oil_and_gas_asset': '油气资产', 'othr_current_liab': '其他流动负债', 'othr_equity_instruments': '其他权益工具', 'othr_payables': '其他应付款', 'othr_receivables': '其他应收款', 'perpetual_bond': '永续债', 'productive_biological_assets': '生产性生物资产', 'project_goods_and_material': '工程物资', 'salable_financial_assets': '可供出售金融资产', 'saleable_finacial_assets': '可供出售金融资产', 'special_payable': '专项应付款', 'special_reserve': '专项储备', 'to_sale_asset': '划分为持有待售的资产', 'to_sale_debt': '划分为持有待售的负债', 'tradable_fnncl_assets': '交易性金融资产', 'tradable_fnncl_liab': '交易性金融负债', 'treasury_stock': '减：库存股', 'unearned_premium_reserve': '未到期责任准备金', 'advance_premium': '预收保费', 'amortized_cost_fnncl_assets': '以摊余成本计量的金融资产', 'assured_pledge_loan': '保户质押贷款', 'assured_saving_and_invest': '保户储金及投资款', 'borrowing_funds': '拆入资金', 'charge_and_commi_payable': '应付手续费及佣金', 'claim_payable': '应付赔付款', 'derivative_fnncl_assets': '衍生金融资产', 'disbursement_loan_and_advance': '发放贷款和垫款', 'dvdnd_payable_for_the_insured': '应付保单红利', 'fixed_deposit': '定期存款', 'fnncl_assets_sold_for_repur': '卖出回购金融资产款', 'fv_chg_income_fnncl_assets': '以公允价值计量且其变动计入其他综合收益的金融资产', 'independent_account_liab': '独立账户负债', 'insurance_contract_reserve': '保险合同准备金', 'interbank_deposit_etc': '同业及其他金融机构存放款项', 'lending_fund': '拆出资金', 'life_insurance_reserve': '寿险责任准备金', 'lt_health_insurance_reserve': '长期健康险责任准备金', 'othr_assets': '其他资产', 'othr_liab': '其他负债', 'paid_capital_deposit': '存出资本保证金', 'premium_receivable': '应收保费', 'receivable_rein_duty_reserve': '应收分保寿险责任准备金', 'received_deposit': '存入保证金', 'rein_account_receivable': '应收分保账款', 'rein_contract_reserve': '应收分保合同准备金', 'rein_payable': '应付分保账款', 'rein_undue_liability_reserve': '应收分保未到期责任准备金', 'separate_account': '独立账户资产'}
    cookies = Get_Cookie(original_url)
    #股票代码及名称字典
    logging.info('正在获取最新的股票列表')
    ipoCode,ipoCodecn,ipoCodehk,ipoCodeus = Get_IPOList(cookies)

    app = QApplication(sys.argv)
        #初始化
    myWin = MyMainForm()
        #将窗口控件显示在屏幕上
    myWin.show()
        #程序运行，sys.exit方法确保程序完整退出
    sys.exit(app.exec_())    